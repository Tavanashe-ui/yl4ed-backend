# app/core/excel_utils.py
"""
Excel utilities for YL4ED Membership Management System.
  - parse_uploaded_excel   : Extract member rows from an .xlsx/.xls upload
  - generate_import_template_excel : Generate a blank .xlsx template for bulk import
"""
import io
import re
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# COLUMN ALIASES  (same set as pdf_utils)
# ──────────────────────────────────────────────
COLUMN_ALIASES: dict[str, list[str]] = {
    "name":                     ["name", "first name", "firstname", "given name"],
    "surname":                  ["surname", "last name", "lastname", "family name"],
    "date_of_birth":            ["date of birth", "dob", "birth date", "birthdate"],
    "gender":                   ["gender", "sex"],
    "phone_number":             ["phone", "phone number", "mobile", "contact number", "cell"],
    "national_identity_number": ["national id", "national identity number", "id number",
                                 "national id number", "id no", "nin"],
    "email_address":            ["email", "email address", "e-mail"],
    "residential_address":      ["address", "residential address", "home address"],
    "occupation":               ["occupation", "job", "profession", "work"],
    "place_of_birth":           ["place of birth", "birth place", "birthplace", "pob"],
    "province_name":            ["province", "province name"],
    "district_name":            ["district", "district name", "current district"],
}

REQUIRED = [
    "name", "surname", "date_of_birth", "gender",
    "phone_number", "national_identity_number",
    "residential_address", "place_of_birth",
    "province_name", "district_name",
]


def _normalise(raw: str) -> str:
    return (raw or "").strip().lower()


def _map_headers(headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(headers):
        # Remove any trailing asterisk and optional spaces before it
        cleaned = re.sub(r'\s*\*\s*$', '', raw)
        norm = _normalise(cleaned)
        for field, aliases in COLUMN_ALIASES.items():
            if norm in aliases:
                mapping[idx] = field
                break
    return mapping


def _parse_date(raw: str) -> Optional[str]:
    """Try multiple formats, return YYYY-MM-DD or None."""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y",
                "%d %B %Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _cell_str(cell_value) -> str:
    """Convert any openpyxl cell value to a clean string."""
    if cell_value is None:
        return ""
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    return str(cell_value).strip()


def _validate_row(data: dict[str, str]) -> list[str]:
    issues = []
    for field in REQUIRED:
        if not data.get(field, "").strip():
            issues.append(f"Missing required field: '{field}'")
    gender_raw = data.get("gender", "").strip().upper()
    if gender_raw and gender_raw not in ("M", "F", "MALE", "FEMALE"):
        issues.append(f"Invalid gender value: '{data.get('gender')}' (expected M or F)")
    dob = data.get("date_of_birth", "").strip()
    if dob and not _parse_date(dob):
        issues.append(f"Unrecognised date format for date_of_birth: '{dob}'")
    return issues


def _clean_row(data: dict[str, str]) -> dict[str, str]:
    cleaned = dict(data)
    g = cleaned.get("gender", "").strip().upper()
    cleaned["gender"] = "M" if g in ("M", "MALE") else "F" if g in ("F", "FEMALE") else g
    dob = cleaned.get("date_of_birth", "")
    parsed = _parse_date(dob)
    if parsed:
        cleaned["date_of_birth"] = parsed
    return cleaned


# ──────────────────────────────────────────────
# IMPORT: PARSE UPLOADED EXCEL
# ──────────────────────────────────────────────
def parse_uploaded_excel(file_bytes: bytes) -> dict:
    """
    Parse a bulk-import .xlsx file.

    Returns:
        {
          "records":      [ {field: value, ...}, ... ],
          "errors":       [ {"row": int, "data": dict, "issues": [str]}, ... ],
          "total_found":  int,
          "valid_count":  int,
          "error_count":  int,
        }
    """
    records: list[dict] = []
    errors:  list[dict] = []

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active  # use first sheet

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return {"records": [], "errors": [], "total_found": 0,
                "valid_count": 0, "error_count": 0}

    # First row = headers
    raw_headers = [_cell_str(h) for h in rows[0]]
    col_map = _map_headers(raw_headers)

    if len(col_map) < 4:
        return {"records": [], "errors": [], "total_found": 0,
                "valid_count": 0, "error_count": 0}

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_data: dict[str, str] = {}
        for col_idx, field_name in col_map.items():
            cell_val = row[col_idx] if col_idx < len(row) else None
            row_data[field_name] = _cell_str(cell_val)

        issues = _validate_row(row_data)
        if issues:
            errors.append({"row": row_idx, "data": row_data, "issues": issues})
        else:
            records.append(_clean_row(row_data))

    return {
        "records":     records,
        "errors":      errors,
        "total_found": len(records) + len(errors),
        "valid_count": len(records),
        "error_count": len(errors),
    }


# ──────────────────────────────────────────────
# TEMPLATE: GENERATE BLANK .xlsx
# ──────────────────────────────────────────────
def generate_import_template_excel() -> io.BytesIO:
    """
    Generate a blank, formatted .xlsx import template with:
    - Branded header row (navy fill, white bold text)
    - Column headers matching COLUMN_ALIASES keys
    - 50 blank data rows with alternating light fill
    - Frozen header row
    - Dropdowns for Gender column
    - Column width hints
    """
    NAVY   = "003580"
    GOLD   = "C9A84C"
    LIGHT  = "EEF2FF"
    WHITE  = "FFFFFF"
    MID    = "F3F4F6"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    # ── Column definitions ──
    columns = [
        ("Name",                     "name",                     20),
        ("Surname",                  "surname",                  20),
        ("Date of Birth",            "date_of_birth",            18),
        ("Gender",                   "gender",                   10),
        ("Phone Number",             "phone_number",             18),
        ("National ID",              "national_identity_number", 22),
        ("Email",                    "email_address",            28),
        ("Address",                  "residential_address",      35),
        ("Place of Birth",           "place_of_birth",           20),
        ("Province",                 "province_name",            20),
        ("District",                 "district_name",            20),
        ("Occupation",               "occupation",               20),
    ]

    REQUIRED_COLS = {c[1] for c in columns if c[1] in REQUIRED}

    # ── Header row styles ──
    header_fill   = PatternFill("solid", fgColor=NAVY)
    header_font   = Font(bold=True, color=WHITE, size=10)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gold_bottom   = Border(bottom=Side(style="medium", color=GOLD))

    # ── Data row styles ──
    even_fill = PatternFill("solid", fgColor=LIGHT)
    odd_fill  = PatternFill("solid", fgColor=WHITE)
    data_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Write header row ──
    ws.row_dimensions[1].height = 32
    for col_idx, (label, field, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        # Mark required fields with *
        cell.value = f"{label} *" if field in REQUIRED_COLS else label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = gold_bottom
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Write 50 blank data rows ──
    for row_idx in range(2, 52):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = data_align

    # ── Gender dropdown (column D = index 4) ──
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"M,F"', allow_blank=True,
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Invalid Gender",
                        error="Please enter M (Male) or F (Female).")
    dv.sqref = f"D2:D51"
    ws.add_data_validation(dv)

    # ── Instructions sheet ──
    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions["A"].width = 80
    instructions = [
        ("YL4ED Bulk Member Import Template", True),
        ("", False),
        ("INSTRUCTIONS:", True),
        ("1. Fill in member data starting from row 2 of the 'Members' sheet.", False),
        ("2. Fields marked with * are required — do not leave them blank.", False),
        ("3. Date of Birth format: YYYY-MM-DD  (e.g. 1998-05-20)", False),
        ("4. Gender: enter M for Male or F for Female only.", False),
        ("5. Province and District must match names registered in the system exactly.", False),
        ("6. Save the file as .xlsx and upload via the Bulk Operations page.", False),
        ("", False),
        ("COLUMN REFERENCE:", True),
    ]
    for field, (label, _, __) in zip(COLUMN_ALIASES.keys(), columns):
        req = " (REQUIRED)" if field in REQUIRED_COLS else " (optional)"
        instructions.append((f"  • {label}{req}", False))

    for r_idx, (text, bold) in enumerate(instructions, start=1):
        cell = ws_info.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10,
                         color=NAVY if bold else "374151")

    # ── Freeze header ──
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
